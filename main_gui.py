# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import os
import sys
import glob


# ========= 核心处理逻辑 =========
def clean(val):
    if val is None:
        return None
    return str(val).strip()


def row_has_data(sheet, row):
    for col in range(3, 34):
        if clean(sheet.cell(row=row, column=col).value):
            return True
    return False


def process_file(file_path, log_callback=None):
    try:
        wb = load_workbook(file_path, data_only=True)

        sheet_att = wb["考勤"]
        sheet_map = wb["工时对照表"]
        sheet_out = wb["拆分工时"]

        # 工时对照表
        map_headers = []
        for col in range(10, 18):
            map_headers.append(clean(sheet_map.cell(row=2, column=col).value))

        mapping = {}
        for row in range(3, sheet_map.max_row + 1):
            emp_id = clean(sheet_map.cell(row=row, column=2).value)
            if not emp_id:
                continue

            values = []
            for col in range(10, 18):
                v = sheet_map.cell(row=row, column=col).value
                values.append(v if isinstance(v, (int, float)) else 0)

            mapping[emp_id] = dict(zip(map_headers, values))

        # 输出表头
        out_headers = []
        for col in range(5, 13):
            out_headers.append(clean(sheet_out.cell(row=4, column=col).value))

        current_row = 8

        while True:
            if not row_has_data(sheet_att, current_row):
                break

            result = {h: 0 for h in out_headers}

            for col in range(3, 34):
                emp_id = clean(sheet_att.cell(row=current_row, column=col).value)
                if not emp_id or emp_id not in mapping:
                    continue

                emp_data = mapping[emp_id]

                for h in out_headers:
                    if h in emp_data:
                        result[h] += emp_data[h]

            out_row = current_row + 1

            for idx, h in enumerate(out_headers):
                sheet_out.cell(row=out_row, column=5 + idx).value = result[h]

            if log_callback:
                log_callback(f"{os.path.basename(file_path)} 第 {current_row} 行完成")

            current_row += 1

        # 输出文件名
        new_file = file_path.replace(".xlsx", "_结果.xlsx")
        wb.save(new_file)

        if log_callback:
            log_callback(f"✅ 完成：{new_file}")

    except Exception as e:
        if log_callback:
            log_callback(f"❌ 失败：{file_path} -> {str(e)}")


# ========= GUI =========
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("工时统计工具")
        self.root.geometry("500x400")

        tk.Button(root, text="选择Excel文件", command=self.select_files, height=2).pack(pady=10)
        tk.Button(root, text="选择文件夹（批量）", command=self.select_folder, height=2).pack(pady=10)

        self.log = tk.Text(root, height=15)
        self.log.pack(fill="both", expand=True)

    def write_log(self, msg):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.root.update()

    def select_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel", "*.xlsx")])
        for f in files:
            process_file(f, self.write_log)

        messagebox.showinfo("完成", "处理完成！")

    def select_folder(self):
        folder = filedialog.askdirectory()
        files = glob.glob(os.path.join(folder, "*.xlsx"))

        for f in files:
            if "_结果" in f:
                continue
            process_file(f, self.write_log)

        messagebox.showinfo("完成", "批量处理完成！")


# ========= 拖拽支持 =========
def handle_drag_file():
    if len(sys.argv) > 1:
        for file_path in sys.argv[1:]:
            if file_path.endswith(".xlsx"):
                process_file(file_path, print)
        input("处理完成，按回车退出...")
        sys.exit()


# ========= 主入口 =========
if __name__ == "__main__":
    # 拖拽运行
    handle_drag_file()

    # GUI运行
    root = tk.Tk()
    app = App(root)
    root.mainloop()
